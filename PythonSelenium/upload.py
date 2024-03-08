#from selenium.webdriver.chrome import webdriver
import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

def update_excel_data(file_path, searchTerm, colName,new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    # to get column value
    for i in range(1, sheet.max_column+ 1):
        if sheet.cell(row=1, column=i).value == colName:
         Dict["col"] =i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == searchTerm:
             Dict["row"] =i

    # Edit the excel with updated value
    sheet.cell(row =Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:\\Users\\Abhishek\\Desktop\\MY\\Learning\\Python\\Excercise\\Upload\\download.xlsx"
fruit_name = "Kivi"
new_Value = "999"
service_obj = Service("C:\\Users\Abhishek\Desktop\MY\Learning\Python\Drivers\chromedriver-win64\chromedriver.exe")
driver = webdriver.Chrome(service=service_obj)

driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

# Edit the excel with updated value
update_excel_data(file_path,fruit_name, "price", new_Value)

#upload file
file_input = driver.find_element(By.XPATH, "//input[@type= 'file']")
file_input.send_keys(file_path)


toast_locator=(By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
WebDriverWait(driver, 5).until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH, "//div[text()= 'Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div//div[@id='cell-"+price_column+"-undefined']").text

print(actual_price)
print(new_Value)
assert actual_price == new_Value